from openpyxl import Workbook
import openpyxl
from scrap import *
from datetime import datetime
from pathlib import Path

def sheetGeneration(filename):

    file = f'csv/{filename}'
    wb = openpyxl.load_workbook(file)

    ws = wb['Sheet1']


    names = 'Not Found'
    for index, row in enumerate(ws.rows,start=1):
        if len(row) > 1:
            payload = {'idno': row[0].value, 'yob': row[1].value}
            data = execute(payload)
                        
        ws.cell(row=index,column=3).value=data['msg']
        if data['data']:
            for i, val in enumerate(data['data']):
                ws.cell(row=index,column=4+i).value=val


    


    file_gen = f'csv/result-doc-{datetime.now()}.xlsx'

    wb.save(filename=file_gen)

    return Path(file_gen)